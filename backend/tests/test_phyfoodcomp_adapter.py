"""Tests for app.phyfoodcomp_adapter — parses a real PhyFoodComp_1.0.xlsx
layout (confirmed by inspecting the actual downloaded file; see
docs/phytate-evidence-review.md) into RawObservation rows. Builds a small
synthetic workbook matching that real structure (two-row header, one row
per food entry, tagname-suffixed content columns alongside ratio columns
that must be ignored) rather than shipping the real file — see
ingest_phytate.py's module docstring on why no real PhyFoodComp data
lives in this repo."""

import openpyxl
import pytest

from app.ingest_phytate import RawObservation
from app.phyfoodcomp_adapter import _decode_preparation_state, load_phyfoodcomp_workbook

HEADER = (
    "Food item ID", "Old code (as in the original source)", "Food Group", "Subgroup",
    "FoodEx2 CODE", "FoodEx2 NAME", "Missing facet", "Exact match", "Matching comments",
    "Country, region", "Type          ", "Food name in own language", "Food name in English",
    "Processing / Influencing factors", "Species/Subspecies", "Cultivar/Variety/Accession Name",
    "Season", "Other", "n", "Comments on data processing/methods", "Publication year",
    "Biblioid ", "Compiler ID", "Latest Revision in Version", "Analytical/Biodiversity",
    "Comments on why some data is not entered", "WATER(g)", "FE(mg)", "ZN(mg)", "CA(mg)",
    "PHYTCPPI(mg)", "PHYTCPPD(mg)", "PHYTCPP(mg)", "PHYTCA (mg)", "PHYTC-(mg)", "PPI(mg)",
    "PPD(mg)", "PP-(mg)", "XP", "IP3(mg)", "IP4(mg)", "IP5(mg)", "IP6(mg)", "IP5_A_IP6 (mg)",
    "IP4_A_IP5_A_IP6(mg)", "IPSUM (mg)", "PHYT-", "PHYTCPPI:FE", "PHYTCPPI:ZN",
)
LEGEND_ROW = (None,) * len(HEADER)


def _col(name: str) -> int:
    return HEADER.index(name)


def _blank_row() -> list:
    return [None] * len(HEADER)


@pytest.fixture
def workbook_path(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cereals = wb.create_sheet("01 Cereals and their products")
    cereals.append(HEADER)
    cereals.append(LEGEND_ROW)

    row1 = _blank_row()
    row1[_col("Food item ID")] = "01010004"
    row1[_col("Food name in English")] = "Bread, brown rice-based, fermented"
    row1[_col("Processing / Influencing factors")] = "p/f/bk"
    row1[_col("PHYTCPPI(mg)")] = 547.507
    row1[_col("IP6(mg)")] = 72.6
    row1[_col("PHYTCPPI:FE")] = 6.58  # ratio column — must be ignored
    cereals.append(row1)

    # a row with a below-detection-limit marker instead of a number
    row2 = _blank_row()
    row2[_col("Food item ID")] = "01010006"
    row2[_col("Food name in English")] = "Breakfast cereal, puffed rice"
    row2[_col("IP3(mg)")] = "< LOD"
    row2[_col("IP4(mg)")] = 40.0
    cereals.append(row2)

    # a row with no usable food description at all
    row3 = _blank_row()
    row3[_col("Food item ID")] = "01010099"
    row3[_col("PHYTCPPI(mg)")] = 100.0
    cereals.append(row3)

    # a row with no phytate-content values at all (only mineral columns)
    row4 = _blank_row()
    row4[_col("Food item ID")] = "01010100"
    row4[_col("Food name in English")] = "Some cereal with no phytate data"
    row4[_col("FE(mg)")] = 2.3
    cereals.append(row4)

    # a non-data reference sheet — must be skipped entirely
    intro = wb.create_sheet("Introductionand and copyright")
    intro.append(["Not a data row"])

    path = tmp_path / "PhyFoodComp_1.0_test.xlsx"
    wb.save(path)
    return path


def test_decode_preparation_state_maps_known_codes():
    assert _decode_preparation_state("p/f/bk") == "processed fermented baked"


def test_decode_preparation_state_keeps_unknown_codes_verbatim():
    assert _decode_preparation_state("zz") == "zz"


def test_decode_preparation_state_none_for_blank():
    assert _decode_preparation_state(None) is None
    assert _decode_preparation_state("") is None


def test_load_phyfoodcomp_workbook_builds_one_observation_per_populated_tagname(workbook_path):
    observations, stats = load_phyfoodcomp_workbook(workbook_path)

    row1_obs = [o for o in observations if o.row_identifier and o.row_identifier.startswith("01010004:")]
    assert len(row1_obs) == 2
    fractions = {o.compound_fraction for o in row1_obs}
    assert fractions == {"PHYTCPPI", "IP6"}

    phytcppi = next(o for o in row1_obs if o.compound_fraction == "PHYTCPPI")
    assert phytcppi.value == 547.507
    assert phytcppi.unit == "mg"
    assert phytcppi.basis == "per_100g_edible_portion"
    assert phytcppi.food_description == "Bread, brown rice-based, fermented"
    assert phytcppi.preparation_state == "processed fermented baked"
    assert phytcppi.analytical_method == "Phytic acid, determined by indirect precipitation"


def test_ratio_columns_are_never_read_as_observations(workbook_path):
    observations, _ = load_phyfoodcomp_workbook(workbook_path)
    assert all(o.compound_fraction != "PHYTCPPI:FE" for o in observations)
    assert all(":" not in o.compound_fraction for o in observations)


def test_non_numeric_cell_is_preserved_not_skipped_and_counted(workbook_path):
    """Prompt 5 (prompts.txt): a censored cell becomes its own
    RawObservation with value=None, not a dropped/skipped row."""
    observations, stats = load_phyfoodcomp_workbook(workbook_path)
    row2_obs = [o for o in observations if o.row_identifier and o.row_identifier.startswith("01010006:")]
    assert {o.compound_fraction for o in row2_obs} == {"IP3", "IP4"}
    assert stats["censored_observations_built"] == 1

    ip3 = next(o for o in row2_obs if o.compound_fraction == "IP3")
    assert ip3.value is None
    assert ip3.value_text == "< LOD"
    assert ip3.value_qualifier == "below_detection_limit"
    assert ip3.unit == "mg"  # unit is still known even though the value is censored

    ip4 = next(o for o in row2_obs if o.compound_fraction == "IP4")
    assert ip4.value == 40.0
    assert ip4.value_qualifier == "measured"


def test_row_with_no_description_is_skipped_and_counted(workbook_path):
    observations, stats = load_phyfoodcomp_workbook(workbook_path)
    assert all(not (o.row_identifier or "").startswith("01010099:") for o in observations)
    assert stats["rows_skipped_no_description"] == 1


def test_row_with_no_phytate_columns_produces_no_observations(workbook_path):
    observations, _ = load_phyfoodcomp_workbook(workbook_path)
    assert all(not (o.row_identifier or "").startswith("01010100:") for o in observations)


def test_non_data_reference_sheets_are_skipped(workbook_path):
    _, stats = load_phyfoodcomp_workbook(workbook_path)
    assert stats["sheets"] == 1  # only "01 Cereals and their products", not "Introductionand and copyright"


def test_every_observation_is_a_raw_observation(workbook_path):
    observations, _ = load_phyfoodcomp_workbook(workbook_path)
    assert observations
    assert all(isinstance(o, RawObservation) for o in observations)


def test_stats_counts_match_observations_built(workbook_path):
    observations, stats = load_phyfoodcomp_workbook(workbook_path)
    assert stats["observations_built"] == len(observations)
    assert stats["rows_considered"] == 4


# ---- Prompt 5 (prompts.txt): censored/non-numeric value preservation ----
# Each test builds its own tiny single-row workbook (rather than adding
# more rows to `workbook_path`, whose exact row/observation counts other
# tests above assert on) covering one value_qualifier at a time.

def _single_cell_workbook(tmp_path, cell_value):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet = wb.create_sheet("01 Cereals and their products")
    sheet.append(HEADER)
    sheet.append(LEGEND_ROW)
    row = _blank_row()
    row[_col("Food item ID")] = "1"
    row[_col("Food name in English")] = "Test food"
    row[_col("IP6(mg)")] = cell_value
    sheet.append(row)
    path = tmp_path / "single_cell.xlsx"
    wb.save(path)
    return path


@pytest.mark.parametrize("cell_value,expected_qualifier", [
    (250.0, "measured"),
    (0.0, "reported_zero"),
    ("< LOD", "below_detection_limit"),
    ("< LOQ", "below_quantification_limit"),
    ("trace", "trace"),
    ("tr", "trace"),
    ("n/a", "unparseable"),
])
def test_value_qualifier_classification(tmp_path, cell_value, expected_qualifier):
    observations, _ = load_phyfoodcomp_workbook(_single_cell_workbook(tmp_path, cell_value))
    assert len(observations) == 1
    assert observations[0].value_qualifier == expected_qualifier


def test_literal_zero_is_not_confused_with_below_detection_limit(tmp_path):
    observations, _ = load_phyfoodcomp_workbook(_single_cell_workbook(tmp_path, 0.0))
    obs = observations[0]
    assert obs.value == 0.0
    assert obs.value_qualifier == "reported_zero"
    # openpyxl round-trips a whole-number Excel cell as a Python int, not
    # a float -- "0", not "0.0". This is genuinely what the source cell
    # itself contains once written to .xlsx, not a bug in this adapter.
    assert obs.value_text == "0"


@pytest.mark.parametrize("cell_value", ["< LOD", "< LOQ", "trace", "n/a"])
def test_censored_values_never_fabricate_a_number(tmp_path, cell_value):
    observations, _ = load_phyfoodcomp_workbook(_single_cell_workbook(tmp_path, cell_value))
    assert observations[0].value is None


def test_censored_value_never_invents_a_detection_limit(tmp_path):
    """PhyFoodComp's own '< LOD' cells carry no accompanying numeric
    limit -- this adapter must not invent one."""
    observations, _ = load_phyfoodcomp_workbook(_single_cell_workbook(tmp_path, "< LOD"))
    assert observations[0].detection_limit_value is None
    assert observations[0].detection_limit_unit is None


def test_blank_string_cell_is_skipped_like_a_none_cell(tmp_path):
    """A cell holding an empty string (as opposed to no entry at all,
    openpyxl's None) carries no signal either -- skipped, not preserved
    as an 'unparseable' observation."""
    observations, stats = load_phyfoodcomp_workbook(_single_cell_workbook(tmp_path, "   "))
    assert observations == []


def test_original_cell_text_round_trips_exactly(tmp_path):
    observations, _ = load_phyfoodcomp_workbook(_single_cell_workbook(tmp_path, "< LOD"))
    assert observations[0].value_text == "< LOD"


def test_unit_and_basis_are_still_present_for_a_censored_observation(tmp_path):
    """Required by prompts.txt PROMPT 5's acceptance test list ('missing
    unit/basis') -- for PhyFoodComp specifically, unit/basis are always
    known per-tagname even when the value itself is censored; this
    confirms that isn't accidentally lost alongside the value."""
    observations, _ = load_phyfoodcomp_workbook(_single_cell_workbook(tmp_path, "< LOD"))
    obs = observations[0]
    assert obs.unit == "mg"
    assert obs.basis == "per_100g_edible_portion"
