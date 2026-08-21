"""Real column adapter for FAO/INFOODS/IZiNCG PhyFoodComp1.0
(`PhyFoodComp_1.0.xlsx`), converting its actual layout into the
RawObservation shape ingest_phytate.py's classification/idempotency
logic already expects — the piece that ingest_phytate.py's module
docstring said would be needed once the real file was in hand (it now
is; see docs/phytate-evidence-review.md §1 for how it was obtained and
its confirmed non-commercial-only licence).

Workbook structure (confirmed by opening the real file, not assumed):
19 food-group data sheets ("01 Cereals and their products" ... "19
Complex recipes") plus "Introduction and copyright", "Codes",
"Components", "Food groups", and "Bibliography" reference sheets. Every
data sheet shares the same two-row header (a label row, then a
legend/definitions row) followed by one row per food entry. Column
*names*, not positions, are relied on below — sheets differ slightly in
which columns they carry (e.g. some food groups have no
Species/Cultivar columns), and PhyFoodComp is generic across food
groups by construction, not something this adapter should assume a
fixed layout for.

Per the Components sheet, every compound-content column's denominator is
"/100g EP" (edible portion, fresh-weight-equivalent) — PhyFoodComp
normalises to this at compilation time, converting from dry-matter where
the source study reported it that way (see docs/phytate-evidence-review.md
§2). That means the number in this file already IS what this app treats
as "the source's published value" (PhyFoodComp is the cited source, not
the underlying primary study it compiled from) — original_basis is always
per_100g_edible_portion here, and there is nothing further for this app to
normalise, so normalised_value/basis/method stay unset.

Only compound-CONTENT columns are read (phytic acid by various analytical
methods, phytate-phosphorus, individual/summed inositol phosphates) — the
pre-computed phytate:mineral ratio columns and the WATER/FE/ZN/CA mineral
content columns are out of scope for a `compound="phytate"` ingestion and
are not read here. A single food-entry row commonly reports more than one
of these columns (e.g. both an indirect-precipitation phytic-acid value
and a directly-measured IP6 value) — each populated column becomes its
own RawObservation, sharing that row's food description/matching fields
but carrying its own compound_fraction/analytical_method/value.

A cell holding a non-numeric marker (e.g. "< LOD", below the analytical
method's limit of detection) is preserved as its own RawObservation with
value=None and an honest value_qualifier (Prompt 5, prompts.txt) — never
coerced into a fabricated number, and never silently dropped either, so
the same food/compound_fraction/row_identifier provenance a real
measured value gets is also available for a censored one. A genuinely
blank cell (openpyxl gives None, meaning the source simply has no entry
in that column at all — not even a marker) is still skipped: there is no
signal to preserve there, unlike a cell that positively states "< LOD"
or similar.
"""

from pathlib import Path

import openpyxl

from .ingest_phytate import RawObservation

# Substring markers -> value_qualifier, checked in this order (most
# specific first) against the lowercased, stripped cell text. "loq"/
# "quantification" must be checked before the generic "<"/"lod" check,
# since a real source cell could plausibly read "< LOQ" (which also
# contains "<"). Never invents a detection_limit_value/unit here (see
# CompoundObservation.detection_limit_value's column comment) — PhyFoodComp's
# own cells carry no numeric limit alongside these markers, only the
# marker text itself.
_CENSORED_MARKERS: tuple[tuple[str, str], ...] = (
    ("loq", "below_quantification_limit"),
    ("quantification", "below_quantification_limit"),
    ("lod", "below_detection_limit"),
    ("detection", "below_detection_limit"),
    ("trace", "trace"),
)
# Exact (not substring) aliases -- "tr" is INFOODS' own short form for
# trace, but is too short to safely match as a substring of arbitrary
# other text.
_CENSORED_EXACT_ALIASES: dict[str, str] = {"tr": "trace"}


def _classify_censored_cell(cell_text: str) -> str:
    """cell_text is already known non-empty and non-numeric by the time
    this is called (see load_phyfoodcomp_workbook) -- classifies which of
    Prompt 5's seven value_qualifier values it represents."""
    lowered = cell_text.strip().lower()
    if lowered in _CENSORED_EXACT_ALIASES:
        return _CENSORED_EXACT_ALIASES[lowered]
    for marker, qualifier in _CENSORED_MARKERS:
        if marker in lowered:
            return qualifier
    if lowered.startswith("<"):
        # An unrecognised "< ..." marker -- still clearly a
        # below-some-limit claim, and detection limits are the far more
        # common convention in food-composition tables than
        # quantification limits, so this is the more conservative
        # (narrower-claim) of the two "below" qualifiers when the source
        # text doesn't say which.
        return "below_detection_limit"
    return "unparseable"


# food-group data sheets only — the reference/index sheets (Introduction,
# Codes, Components, Food groups, Bibliography) have a different layout
# and carry no food-entry rows.
_NON_DATA_SHEETS = {
    "Introductionand and copyright", "Codes", "Components", "Food groups", "Bibliography",
}

# INFOODS tagname -> (unit, human-readable analytical method), from the
# Components sheet. Deliberately excludes WATER/FE/ZN/CA (mineral/water
# content, not phytate) and every ":"-containing ratio column (a derived
# figure, not a compound-content observation this table's scope covers —
# see prompts.txt's explicit non-goal on absorption/derived modelling).
_PHYTATE_TAGNAMES: dict[str, tuple[str, str]] = {
    "PHYTCPPI": ("mg", "Phytic acid, determined by indirect precipitation"),
    "PHYTCPPD": ("mg", "Phytic acid, determined by direct precipitation"),
    "PHYTCPP": ("mg", "Phytic acid, determined by anion exchange"),
    "PHYTCA": ("mg", "Phytic acid, determined by colorimetry after alkaline phosphatase hydrolyzation (K-PHYT kit)"),
    "PHYTC-": ("mg", "Phytic acid, determined by colorimetry (method unspecified in source)"),
    "PPI": ("mg", "Phytate phosphorus, determined by indirect precipitation"),
    "PPD": ("mg", "Phytate phosphorus, determined by direct precipitation"),
    "PP-": ("mg", "Phytate phosphorus, method unspecified in source"),
    "IP3": ("mg", "Inositol triphosphate (HPLC/HPAE)"),
    "IP4": ("mg", "Inositol tetraphosphate (HPLC/HPAE)"),
    "IP5": ("mg", "Inositol pentaphosphate (HPLC/HPAE)"),
    "IP6": ("mg", "Inositol hexaphosphate (HPLC/HPAE)"),
    "IP5_A_IP6": ("mg", "Inositol penta- + hexaphosphate, summed (HPLC/HPAE)"),
    "IP4_A_IP5_A_IP6": ("mg", "Inositol tetra- + penta- + hexaphosphate, summed (HPLC/HPAE)"),
    "IPSUM": ("mg", "Total inositol phosphates, summed (HPLC/HPAE)"),
    "PHYT-": ("mg", "Phytic acid, method unknown or variable"),
}

_NAMED_COLUMNS = (
    "Food item ID", "Food name in English", "Food name in own language", "FoodEx2 NAME",
    "Processing / Influencing factors",
)

# From the "Processing / Influencing factors" column legend, repeated at
# the top of every data sheet — decodes e.g. "p/f/bk" into "processed
# fermented baked". An unmapped segment (a code not in this table) is
# kept verbatim rather than dropped, so an update to PhyFoodComp's own
# code list doesn't silently lose information.
_PROCESSING_CODES: dict[str, str] = {
    "r": "raw", "d": "dried", "p": "processed", "s-": "soaked", "ws": "water-soaked",
    "wss": "water and salt-soaked", "as": "ash-soaked", "f": "fermented", "f-": "fermented",
    "pb": "parboiled", "wb": "water-boiled", "wsb": "water and salt-boiled", "b-": "boiled",
    "bk": "baked", "rp": "recipe", "rpi": "industrial recipe", "ro": "roasted", "st": "steamed",
    "fr": "fried", "ac": "autoclaved", "mw": "microwaved", "c": "cooked", "bl": "blanched",
    "t": "toasted", "sk": "smoked", "cn": "canned", "dt": "defatted", "ex": "extruded",
    "ir": "irradiated", "fz": "frozen", "a": "abrased", "th": "thermally treated", "dh": "dehulled",
    "g": "germinated", "sd": "stored", "ft": "soil fertilization", "pt": "pesticide application",
    "ly": "lyophilization", "gr": "grilled", "ch": "chemically treated",
}


def _decode_preparation_state(raw_code: str | None) -> str | None:
    if not raw_code or not raw_code.strip():
        return None
    segments = raw_code.strip().split("/")
    return " ".join(_PROCESSING_CODES.get(seg.strip().lower(), seg.strip()) for seg in segments if seg.strip())


def _find_column(header: tuple, name: str) -> int | None:
    for i, cell in enumerate(header):
        if cell is not None and str(cell).strip() == name:
            return i
    return None


def _find_tagname_column(header: tuple, tagname: str) -> int | None:
    """A content column's header is the tagname plus a unit suffix, e.g.
    "PHYTCPPI(mg)" — matched by exact tagname prefix immediately followed
    by "(" or whitespace, so "PHYTCPP" (a real tagname) doesn't false-
    -match "PHYTCPPI(mg)"/"PHYTCPPD(mg)"."""
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cell_str = str(cell).strip()
        if cell_str == tagname:
            return i
        if cell_str.startswith(tagname) and cell_str[len(tagname):len(tagname) + 1] in (" ", "("):
            return i
    return None


def load_phyfoodcomp_workbook(xlsx_path: Path) -> tuple[list[RawObservation], dict]:
    """Parses every food-group sheet in a real PhyFoodComp_1.0.xlsx into
    RawObservation rows (compound_fraction=INFOODS tagname, one per
    populated phytate-content cell, numeric or censored) plus stats on
    what was skipped/built and why — rows_considered,
    rows_skipped_no_description, censored_observations_built (e.g.
    "< LOD" cells — now preserved as their own RawObservation with
    value=None, not skipped; see module docstring), observations_built
    (numeric and censored together)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    stats = {"sheets": 0, "rows_considered": 0, "rows_skipped_no_description": 0,
              "censored_observations_built": 0, "observations_built": 0}
    observations: list[RawObservation] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in _NON_DATA_SHEETS:
            continue
        stats["sheets"] += 1
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        next(rows, None)  # legend/definitions row — not data
        if header is None:
            continue

        col = {name: _find_column(header, name) for name in _NAMED_COLUMNS}
        tagname_cols = {
            tagname: _find_tagname_column(header, tagname) for tagname in _PHYTATE_TAGNAMES
        }

        for row in rows:
            if row is None or all(v is None for v in row):
                continue
            stats["rows_considered"] += 1

            def get(key: str):
                idx = col.get(key)
                return row[idx] if idx is not None and idx < len(row) else None

            description = get("Food name in English") or get("Food name in own language") or get("FoodEx2 NAME")
            if not description or not str(description).strip():
                stats["rows_skipped_no_description"] += 1
                continue
            description = str(description).strip()

            row_id = get("Food item ID")
            row_identifier = str(row_id).strip() if row_id is not None and str(row_id).strip() else None
            preparation_state = _decode_preparation_state(get("Processing / Influencing factors"))

            for tagname, (unit, method) in _PHYTATE_TAGNAMES.items():
                idx = tagname_cols.get(tagname)
                if idx is None or idx >= len(row):
                    continue
                cell = row[idx]
                if cell is None:
                    continue  # no entry in this column at all -- no signal to preserve
                cell_text = str(cell).strip()
                if not cell_text:
                    continue  # a blank string cell -- same as no entry

                try:
                    value = float(cell)
                    value_qualifier = "reported_zero" if value == 0.0 else "measured"
                except (TypeError, ValueError):
                    value = None
                    value_qualifier = _classify_censored_cell(cell_text)
                    stats["censored_observations_built"] += 1

                observations.append(RawObservation(
                    food_description=description,
                    value=value,
                    value_text=cell_text,
                    value_qualifier=value_qualifier,
                    unit=unit,
                    basis="per_100g_edible_portion",
                    preparation_state=preparation_state,
                    compound_fraction=tagname,
                    analytical_method=method,
                    row_identifier=f"{row_identifier}:{tagname}" if row_identifier else None,
                ))
                stats["observations_built"] += 1

    return observations, stats
